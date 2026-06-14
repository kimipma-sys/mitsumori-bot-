from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage
import datetime
import os
import database

def format_currency(amount):
    """金額を見やすくフォーマットする（マイナスの場合は -¥5,000 のようにする）"""
    if amount < 0:
        return f"-¥{abs(int(amount)):,}"
    return f"¥{int(amount):,}"

def generate_excel(user_id, items, project_name, doc_type="estimate"):
    wb = Workbook()
    ws = wb.active
    
    # doc_typeに応じてタイトルとファイル名を切り替え
    title_text = "御見積書"
    file_prefix = "estimate"
    if doc_type == "invoice":
        title_text = "御請求書"
        file_prefix = "invoice"
        
    ws.title = title_text
    
    # Header
    ws['A1'] = title_text
    ws['A1'].font = Font(size=20, bold=True)
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Date
    ws['E2'] = datetime.datetime.now().strftime("%Y年%m月%d日")
    ws['E2'].alignment = Alignment(horizontal='right')
    
    # Project Name (件名)
    ws['A3'] = f"件名: {project_name}"
    ws['A3'].font = Font(size=12, underline='single')
    ws.merge_cells('A3:E3')
    
    # Total Calculation
    total_amount = sum(item['total'] for item in items)
    tax = int(total_amount * 0.1)
    grand_total = total_amount + tax
    
    ws['A5'] = f"{title_text}金額: {format_currency(grand_total)} (税込)"
    ws['A5'].font = Font(size=14, bold=True)
    ws.merge_cells('A5:D5')
    
    # Table Headers
    headers = ["No.", "品名・工事名", "数量", "単価", "金額"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        
    # Table Content
    row_num = 9
    for i, item in enumerate(items, 1):
        ws.cell(row=row_num, column=1, value=i)
        ws.cell(row=row_num, column=2, value=item['item_name'])
        
        qty_display = int(item['quantity']) if item['quantity'].is_integer() else item['quantity']
        unit_display = item.get('unit', '')
        ws.cell(row=row_num, column=3, value=f"{qty_display}{unit_display}")
        
        ws.cell(row=row_num, column=4, value=format_currency(item['price']))
        ws.cell(row=row_num, column=5, value=format_currency(item['total']))
        row_num += 1
        
    # Subtotals
    ws.cell(row=row_num+1, column=4, value="小計").font = Font(bold=True)
    ws.cell(row=row_num+1, column=5, value=format_currency(total_amount))
    
    ws.cell(row=row_num+2, column=4, value="消費税(10%)").font = Font(bold=True)
    ws.cell(row=row_num+2, column=5, value=format_currency(tax))
    
    ws.cell(row=row_num+3, column=4, value="合計").font = Font(bold=True)
    ws.cell(row=row_num+3, column=5, value=format_currency(grand_total))
    
    # インボイス登録番号と振込先口座情報（請求書の場合のみ）
    if doc_type == "invoice":
        current_row = row_num + 6
        invoice_number = os.getenv("INVOICE_NUMBER", "")
        bank_info = os.getenv("BANK_INFO", "")
        
        if invoice_number:
            ws.cell(row=current_row, column=1, value=f"登録番号: {invoice_number}")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 2
            
        if bank_info:
            ws.cell(row=current_row, column=1, value="【お振込先】")
            ws.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            ws.cell(row=current_row, column=1, value=bank_info)
            ws.cell(row=current_row, column=1).alignment = Alignment(wrap_text=True)
            ws.merge_cells(f"A{current_row}:E{current_row}")
            ws.row_dimensions[current_row].height = 45
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    
    # === エビデンスシートの追加処理 ===
    evidence_list = database.get_all_evidence(user_id, project_name)
    if evidence_list:
        ws2 = wb.create_sheet(title="エビデンス")
        ws2.column_dimensions['B'].width = 60
        
        ws2['B2'] = f"■ 添付エビデンス一覧 （件名: {project_name}）"
        ws2['B2'].font = Font(size=16, bold=True)
        
        current_row = 4
        for ev in evidence_list:
            ws2.cell(row=current_row, column=2, value=f"【 {ev['title']} 】").font = Font(bold=True, size=14)
            current_row += 2
            
            try:
                if os.path.exists(ev['filepath']):
                    img = ExcelImage(ev['filepath'])
                    ratio = 400.0 / img.width
                    img.width = 400
                    img.height = int(img.height * ratio)
                    
                    ws2.add_image(img, f"B{current_row}")
                    rows_needed = int(img.height / 15) + 3
                    current_row += rows_needed
                else:
                    ws2.cell(row=current_row, column=2, value="※画像ファイルが見つかりません")
                    current_row += 2
            except Exception as e:
                ws2.cell(row=current_row, column=2, value="※画像の読み込みに失敗しました")
                current_row += 2
    
    # Ensure static directory exists
    os.makedirs('static', exist_ok=True)
    
    # ファイル名にも案件名を入れる
    safe_project_name = project_name.replace('/', '_').replace('\\', '_')
    filename = f"{file_prefix}_{safe_project_name}_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    filepath = os.path.join('static', filename)
    wb.save(filepath)
    
    return filename
